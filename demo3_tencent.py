from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.tmt.v20180321 import tmt_client, models
import json
import time
import openpyxl


# 腾讯翻译函数
def tencent_translate(paragraph):
    dic = {}
    try:
        # 修改appid和appkey为自己对应的api
        cred = credential.Credential("appid", "appkey")
        httpProfile = HttpProfile()
        httpProfile.endpoint = "tmt.tencentcloudapi.com"

        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        client = tmt_client.TmtClient(cred, "ap-shanghai", clientProfile)

        req = models.TextTranslateRequest()
        params = '{"SourceText":"%s","Source":"en","Target":"zh","ProjectId":0}' % paragraph
        req.from_json_string(params)

        resp = client.TextTranslate(req)
        dic = resp.to_json_string()
        dic = json.loads(dic)  # 将字符串转换为字典
        return dic["TargetText"] # 输出：我


    except TencentCloudSDKException as err:
        print(err)
        return 'error'

# 打开 Excel 文件
file_path = 'your_excel.xlsx'  # 请替换为你的文件路径
wb = openpyxl.load_workbook(file_path)

# 选择 "Result 1" 工作表
sheet = wb['Result 1']

# 从第1行开始，读取D列英文并翻译到E列
for row in range(1, sheet.max_row + 1):
    english_text = sheet.cell(row=row, column=4).value  # D列是第4列
    if english_text:  # 如果英文文本不为空
        chinese_translation = tencent_translate(english_text)  # 翻译英文到中文
        print(f"Translating row {row}: {english_text} to {chinese_translation}")
        if chinese_translation == 'error':
            break
        sheet.cell(row=row, column=5, value=chinese_translation)  # 将翻译结果写入 E 列（第5列）
        time.sleep(0.1)  # 每秒翻译5次，避免被封禁，翻译算0.1s
    if row % 2000 == 0:
        wb.save(file_path)
        print('写入了', row, '行')

# 保存修改后的 Excel 文件
wb.save(file_path)
print("Translation complete and saved.")
