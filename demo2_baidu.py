import requests
import random
from hashlib import md5
import openpyxl

# 计算 MD5
def make_md5(s, encoding='utf-8'):
    return md5(s.encode(encoding)).hexdigest()
appid = '' #自己的id
appkey = '' #自己的秘钥
# 构造 API 请求 URL
endpoint = 'http://api.fanyi.baidu.com'
path = '/api/trans/vip/translate'
url = endpoint + path
# 创建一个随机盐值
def generate_salt():
    return random.randint(32768, 65536)

# 百度翻译函数
def baidu_translate(query, from_lang, to_lang):
    salt = generate_salt()
    sign = make_md5(appid + query + str(salt) + appkey)

    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {
        'appid': appid,
        'q': query,
        'from': from_lang,
        'to': to_lang,
        'salt': salt,
        'sign': sign
    }

    # 发送翻译请求
    try:
        r = requests.post(url, params=payload, headers=headers)
        r.raise_for_status()  # 如果请求失败，会抛出异常
        result = r.json()
        if "trans_result" in result:
            return result["trans_result"][0]['dst']
        else:
            print(f"Error in translation: {result}")
            return "error"
    except requests.exceptions.RequestException as e:
        print(f"Request failed: {e}")
        return ""

# 打开 Excel 文件
file_path = 'your_excel.xlsx'  # 请替换为你的文件路径
wb = openpyxl.load_workbook(file_path)
# 选择 "Result 1" 工作表
sheet = wb['Result 1']
# 从第1行开始，读取D列英文并翻译到E列
for row in range(1, sheet.max_row + 1):
    english_text = sheet.cell(row=row, column=4).value  # D列是第4列
    if english_text:  # 如果英文文本不为空
        chinese_translation = baidu_translate(english_text, 'en', 'zh')  # 翻译英文到中文
        print(f"Translating row {row}: {english_text} to {chinese_translation}")
        if chinese_translation == 'error':
            break
        sheet.cell(row=row, column=5, value=chinese_translation)  # 将翻译结果写入 E 列（第5列）
        # time.sleep(0.05)  # 每秒翻译10次，避免被百度封禁
    if row % 2000 == 0:
        wb.save(file_path)
        print('写入了', row, '行')
# 保存修改后的 Excel 文件
wb.save(file_path)
print("Translation complete and saved.")
